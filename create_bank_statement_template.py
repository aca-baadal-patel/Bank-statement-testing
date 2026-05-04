import openpyxl
from openpyxl.styles import Alignment, Font

# Create a new workbook and sheets
workbook = openpyxl.Workbook()
summary_sheet = workbook.active
summary_sheet.title = "Summary"

transactions_sheet = workbook.create_sheet(title="Transactions")
debits_sheet = workbook.create_sheet(title="Debits")

# Setup Summary Sheet
# Borrower/Loan Information section
summary_sheet['A1'] = "Borrower/Loan Information"
summary_sheet['A2'] = "Today's Date"
summary_sheet['B2'] = "2026-05-04"
summary_sheet['A3'] = "Completed By"
summary_sheet['B3'] = "Your Name"
summary_sheet['A4'] = "Borrower Name"
summary_sheet['B4'] = "Borrower Name"
summary_sheet['A5'] = "Employee Name"
summary_sheet['B5'] = "Employee Name"
summary_sheet['A6'] = "Account Number"
summary_sheet['B6'] = "Account Number"
summary_sheet['A7'] = "Loan Number"
summary_sheet['B7'] = "Loan Number"
summary_sheet['A8'] = "Account Type"
summary_sheet['B8'] = "Account Type"

# Large Deposit Flag section
summary_sheet['A10'] = "Large Deposit Flag"
summary_sheet['A11'] = "Average Deposit"
summary_sheet['B11'] = "0"  # Placeholder
summary_sheet['A12'] = "Percentage Amount"
summary_sheet['B12'] = "0"  # Placeholder
summary_sheet['A13'] = "Large Deposit Threshold"
summary_sheet['B13'] = "0"  # Placeholder

# Totals section (with example formulas)
summary_sheet['A15'] = "Totals"
summary_sheet['A16'] = "Total Deposits"
summary_sheet['B16'] = "=SUM(Transactions!C:C)"  # Formula reference to Transactions sheet
summary_sheet['A17'] = "Total Included Deposits"
summary_sheet['B17'] = "=SUMIFS(Transactions!C:C, Transactions!E:E, 'Include')"  # Example formula
summary_sheet['A18'] = "Total Excluded Deposits"
summary_sheet['B18'] = "=SUMIFS(Transactions!C:C, Transactions!E:E, 'Exclude')"  # Example formula

# Transaction Type Breakdown
summary_sheet['A20'] = "Breakdown by Transaction Type"
transaction_types = ["ATM_DEPOSIT", "PAYROLL", "ZELLE", "OTHER", "TRANSFER", "DIRECT_DEPOSIT", "REFUNDS", "UNEMPLOYMENT", "LOAN_DEPOSITS", "TAX_RETURN"]
for i, tx_type in enumerate(transaction_types, start=22):
    summary_sheet[f'A{i}'] = tx_type
    summary_sheet[f'B{i}'] = f'=SUMIFS(Transactions!C:C, Transactions!D:D, "{tx_type}", Transactions!E:E, "Include")'  # Included Total
    summary_sheet[f'C{i}'] = f'=SUMIFS(Transactions!C:C, Transactions!D:D, "{tx_type}", Transactions!E:E, "Exclude")'  # Excluded Total
    summary_sheet[f'D{i}'] = f'=COUNTIFS(Transactions!D:D, "{tx_type}", Transactions!E:E, "Include")'  # Included Count

# Setup Transactions Sheet
transactions_header = ["Date", "Transaction Description", "Amount", "Deposit Category", "Include/Exclude", "Large Deposit Flag"]
transactions_sheet.append(transactions_header)

# Setup Debits Sheet
debits_header = ["Date", "Transaction Description", "Amount", "Include/Exclude"]
debits_sheet.append(debits_header)

# Generate Excel file
file_path = 'bank_statement_template.xlsx'
workbook.save(file_path)
print(f'Workbook saved at {file_path}.')
